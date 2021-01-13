from rdflib import Graph, Namespace, Literal, BNode
from rdflib.namespace import OWL, RDF, RDFS, NamespaceManager, XSD
from SPARQLWrapper import SPARQLWrapper, JSON
import csv, random, string, json, urllib
import os.path
from openpyxl import load_workbook
from elasticsearch import Elasticsearch
from alive_progress import alive_bar
import requests

from pdb import set_trace as st
#exit this with c

RRO = Namespace("https://rdf.ng-london.org.uk/raphael/ontology/")
RRI = Namespace("https://rdf.ng-london.org.uk/raphael/resource/")
CRM = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
NGO = Namespace("https://round4.ng-london.org.uk/ex/lcd/")
AAT = Namespace("http://vocab.getty.edu/page/aat/")
TGN = Namespace("http://vocab.getty.edu/page/tgn/")
WD = Namespace("http://www.wikidata.org/entity/")

g = Graph()
g.parse("inputs/rrr_i_v0.5.xml", format="xml")
g.bind('rro',RRO)
g.bind('rri',RRI)

new_graph = Graph()
new_graph.bind('crm',CRM)
new_graph.bind('ngo',NGO)
new_graph.bind('aat',AAT)
new_graph.bind('tgn',TGN)
new_graph.bind('wd',WD)

#print(len(g)) # prints 2

#print(RDF.type)
#print(RDFS.subClassOf)
#print(getattr(RRO,'RP16.has_height_in_cm'))

def query_graph(graph,subj,pred,obj):
    for s,p,o in graph.triples((subj,pred,obj)):
        print(s,p,o)

def map_property(graph,new_graph,old_property,new_property):
    for x,y,z in graph.triples((None,old_property,None)):
        #graph.remove((x,y,z))
        new_graph.add((x,new_property,z))
    return graph, new_graph

def map_class(graph,new_graph,old_class,new_class):
    for x,y,z in graph.subjects((old_class,None,None)):
        #graph.remove((x,y,z))
        new_graph.add((new_class,y,z))

    for x,y,z in graph.objects((None,None,old_class)):
        #graph.remove((x,y,z))
        new_graph.add((x,y,new_class))

    return graph, new_graph

def triples_to_csv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.csv','w',newline='') as f:
        write = csv.writer(f)

        write.writerow(fields)
        write.writerows(triples)

    print('CSV created!')

def get_property(uri):
    remove_uri = uri.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    final_property = remove_uri.replace('_',' ')
    return final_property

def get_json(url):
    operUrl = urllib.request.urlopen(url)
    if(operUrl.getcode()==200):
       data = operUrl.read()
       json_data = json.loads(data)
    else:
       print("Error receiving data", operUrl.getcode())
    return json_data

def check_pids_csv(input_literal):
    pid_value = 'None'
    if os.path.isfile('outputs/placeholder_pids.csv') == True:
        with open('outputs/placeholder_pids.csv','r') as f:
            reader = csv.DictReader(f, delimiter=',')
            dict_list = []
            for row in reader:
                dict_list.append(row)
            for value in range(len(dict_list)):
                csv_literal = str(dict_list[value]['Literal value'])
                if csv_literal == str(input_literal):
                    pid_value = str(dict_list[value]['Placeholder PID'])
                    break
    return pid_value

def check_pids_elasticsearch(search_term): 
    es = Elasticsearch(
        cloud_id = 'crm-migration:dXMtZWFzdC0xLmF3cy5mb3VuZC5pbyQ3ZjcwZDQ5YWQyNTQ0MWFmOTM1ZDJmYTRmZGYwYmI0MCQ0ZWRlZmYwZjM2YWE0MDA2OWNlNTExYmVlNmI1NjQ4OA==',
        http_auth = ('elastic','QsS2wTPlUsjTQNlrvlQbnkzG'),
    )     

    pid_value = 'None'
    res = es.search(index='pids',body={'query':{'match':{'column1':search_term}}})

    for hit in res['hits']['hits']:
        pid_value = hit['_source']['column2']
    
    return pid_value

def find_old_pid(ng_number):
    #json_data = get_json('https://scientific.ng-london.org.uk/export/ngpidexport_00A.json')
    if ng_number.startswith('https'):
        ng_number = str(ng_number.rsplit('/',1)[-1])
    else:
        ng_number
    old_pid = 'None'
    with open('ngpidexports/ngpidexport_00A.json') as f:
        json_data = json.load(f)
        for x in json_data:
            for y in json_data[x]["objects"]:
                if json_data[x]["objects"][y] == ng_number:
                    old_pid = y
                    break
    return old_pid

def generate_placeholder_PID(input_literal):
    N = 4
    placeholder_PID = ""
    for number in range(N):
        res = ''.join(random.choices(string.ascii_uppercase + string.digits, k = N))
        placeholder_PID += str(res)
        placeholder_PID += '-'
    placeholder_PID = placeholder_PID[:-1]
    #input_list = [input_literal, placeholder_PID]
    #fields = ['Literal value','Placeholder PID']
    existing_pid = check_pids_elasticsearch(input_literal)
    old_pid = find_old_pid(input_literal)

    if existing_pid != 'None':
        return existing_pid
    elif old_pid != 'None':
        return old_pid
        '''
    elif os.path.isfile('outputs/placeholder_pids.csv') == True:      
        with open('outputs/placeholder_pids.csv', 'a', newline='') as f:
            write = csv.writer(f)
            write.writerow(input_list)
    else:
        with open('outputs/placeholder_pids.csv','w',newline='') as f:
            write = csv.writer(f)
            write.writerow(fields)
            write.writerow(input_list)
            '''
    else:
        es = Elasticsearch(
            cloud_id = 'crm-migration:dXMtZWFzdC0xLmF3cy5mb3VuZC5pbyQ3ZjcwZDQ5YWQyNTQ0MWFmOTM1ZDJmYTRmZGYwYmI0MCQ0ZWRlZmYwZjM2YWE0MDA2OWNlNTExYmVlNmI1NjQ4OA==',
            http_auth = ('elastic','QsS2wTPlUsjTQNlrvlQbnkzG'),
        )  

        body = {
                'column1' : input_literal,
                'column2' : placeholder_PID
                }
        es.index(index = 'pids',doc_type='_doc',body=body)
        print('new thing added to elasticsearch')

    return placeholder_PID

def create_PID_from_triple(pid_type, subj):
    if pid_type == 'object':
        pid_name = subj
    else:
        pid_name = pid_type + ' of ' + subj
    output_pid = generate_placeholder_PID(pid_name)

    return output_pid

def find_aat_value(material,material_type):
    material = str(get_property(material))
    wb = load_workbook(filename = 'inputs/NG_Meduim_and_Support_AAT.xlsx', read_only=True)
    if material_type == getattr(RRO,'RP20.has_medium'):
        ws = wb['Medium Material']
    elif material_type == getattr(RRO,'RP32.has_support'):
        ws = wb['Support Materials']
        #temporary placeholder until spreadsheet is fixed
        material = 'poplar'
    for row in ws.iter_rows(values_only=True):
        if material in row:
            aat_number_string = str(row[1])
            aat_number = aat_number_string.replace('aat:','')
            aat_type = row[2]
            return aat_number, aat_type
    wb.close()

def wikidata_query(literal):
    sparql = SPARQLWrapper('https://query.wikidata.org/sparql')
    string_literal = str(literal)

    query = ('''
    SELECT DISTINCT ?year
    WHERE
    {
    ?year  wdt:P31 wd:Q577 .
            ?year rdfs:label ?yearLabel .
    FILTER( str(?yearLabel) = "''' + string_literal + '''" ) .
    SERVICE wikibase:label { bd:serviceParam wikibase:language "[AUTO_LANGUAGE],en". }
    }
    ''')

    sparql.setQuery(query)
    sparql.setReturnFormat(JSON)

    ret = sparql.query().convert()
    st()
    for result in ret["results"]["bindings"]:
        wikidata_value = result["year"]["value"]

    return wikidata_value

def create_title_triples(PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP34.has_title'):
        title_PID = create_PID_from_triple('title', subj)

        new_graph.add((getattr(NGO,PID), CRM.P102_has_title, getattr(NGO,title_PID)))
        new_graph.add((getattr(NGO,title_PID), CRM.P1_is_identified_by, Literal(obj)))
        new_graph.add((getattr(NGO,title_PID), CRM.P2_has_type, CRM.E35_Title))
        new_graph.add((getattr(NGO,title_PID), CRM.P2_has_type, getattr(AAT,'300417209')))
        new_graph.add((getattr(AAT,'300417209'), CRM.P1_is_identified_by, Literal('full title@en')))

    elif pred == getattr(RRO, 'RP31.has_short_title'):
        title_PID = create_PID_from_triple('short title', subj)

        new_graph.add((getattr(NGO,PID), CRM.P102_has_title, getattr(NGO,title_PID)))
        new_graph.add((getattr(NGO,title_PID), CRM.P1_is_identified_by, Literal(obj)))
        new_graph.add((getattr(NGO,title_PID), CRM.P2_has_type, CRM.E35_Title))
        new_graph.add((getattr(NGO,title_PID), CRM.P2_has_type, getattr(AAT,'300417208')))
        new_graph.add((getattr(AAT,'300417208'), CRM.P1_is_identified_by, Literal('brief title@en')))

    return new_graph

def create_medium_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP20.has_medium'):
        medium_PID = create_PID_from_triple('medium', subj)
        aat_number, aat_type = find_aat_value(obj, pred)

        new_graph.add((getattr(NGO,subject_PID), CRM.P45_consists_of, getattr(NGO,medium_PID)))
        new_graph.add((getattr(NGO,medium_PID), CRM.P2_has_type, getattr(AAT,aat_number)))
        new_graph.add(((getattr(AAT,aat_number)), CRM.P1_is_identified_by, Literal(aat_type)))
        new_graph.add((getattr(NGO,medium_PID), CRM.P2_has_type, CRM.E57_Material))
        new_graph.add((getattr(NGO,medium_PID), CRM.P2_has_type, getattr(AAT,'300163343')))
        new_graph.add(((getattr(AAT,'300163343')), CRM.P1_is_identified_by, Literal('media (artists\' material)@en')))

    elif pred == getattr(RRO, 'RP32.has_support'):
        medium_PID = create_PID_from_triple('support material', subj)
        aat_number, aat_type = find_aat_value(obj, pred)

        new_graph.add((getattr(NGO,subject_PID), CRM.P45_consists_of, getattr(NGO,medium_PID)))
        new_graph.add((getattr(NGO,medium_PID), CRM.P2_has_type, getattr(AAT,aat_number)))
        new_graph.add(((getattr(AAT,aat_number)), CRM.P1_is_identified_by, Literal(aat_type)))
        new_graph.add((getattr(NGO,medium_PID), CRM.P2_has_type, CRM.E57_Material))
        new_graph.add((getattr(NGO,medium_PID), CRM.P2_has_type, getattr(AAT,'300014844')))
        new_graph.add(((getattr(AAT,'300014844')), CRM.P1_is_identified_by, Literal('supports (artists\' materials)@en')))

    return new_graph

def create_collection_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP99.is_part_of'):
        collection_PID = create_PID_from_triple('object', subj)

        new_graph.add((getattr(NGO,subject_PID), CRM.P46_forms_part_of, getattr(NGO,collection_PID)))
        new_graph.add((getattr(NGO,collection_PID), CRM.P2_has_type, CRM.E78_Curated_Holding))
        new_graph.add((getattr(NGO,collection_PID), CRM.P1_is_identified_by, Literal(obj)))

    return new_graph

def create_dimension_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP36.has_width_in_cm'):
        dimension_PID = create_PID_from_triple('width', subj)

        new_graph.add((getattr(NGO,subject_PID), CRM.P43_has_dimension, getattr(NGO,dimension_PID)))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P2_has_type, CRM.E54_Dimension))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P90_has_value, Literal(obj)))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P91_has_unit, getattr(AAT,'300379098')))
        new_graph.add((getattr(AAT,'300379098'), CRM.P1_is_identified_by, Literal('centimeters@en')))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P2_has_type, getattr(AAT,'300055647')))
        new_graph.add((getattr(AAT,'300055647'), CRM.P1_is_identified_by, Literal('width@en')))

    elif pred == getattr(RRO, 'RP16.has_height_in_cm'): 
        dimension_PID = create_PID_from_triple('height', subj)

        new_graph.add((getattr(NGO,subject_PID), CRM.P43_has_dimension, getattr(NGO,dimension_PID)))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P2_has_type, CRM.E54_Dimension))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P90_has_value, Literal(obj)))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P91_has_unit, getattr(AAT,'300379098')))
        new_graph.add((getattr(AAT,'300379098'), CRM.P1_is_identified_by, Literal('centimeters@en')))
        new_graph.add((getattr(NGO,dimension_PID), CRM.P2_has_type, getattr(AAT,'300055644')))
        new_graph.add((getattr(AAT,'300055644'), CRM.P1_is_identified_by, Literal('height@en')))

    return new_graph

def create_identifier_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP17.has_identifier'):
        blank_node = BNode()
        new_graph.add((getattr(NGO, subject_PID), CRM.P48_has_preferred_identifier, blank_node))
        new_graph.add((blank_node, CRM.P2_has_type, CRM.E42_Identifier))
        new_graph.add((blank_node, CRM.P2_has_type, getattr(AAT, '300312355')))
        new_graph.add((getattr(AAT, '300312355'), CRM.P1_is_identified_by, Literal('accession numbers@en')))
        new_graph.add((blank_node, CRM.P1_is_identified_by, getattr(NGO, obj)))

    return new_graph

def create_type_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP98.is_in_project_category'):
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, obj))

    return new_graph

def create_time_span_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP209.has_time_span'):
        time_span_PID = create_PID_from_triple('time span', subj)
        wikidata_year = wikidata_query(obj)

        #will need to change this logic if we encounter a time span that isn't a year
        new_graph.add((getattr(NGO, subject_PID), CRM.P4_has_time_span, time_span_PID))
        new_graph.add((time_span_PID, CRM.P2_has_type, CRM.E52_Time_span))
        new_graph.add((time_span_PID, CRM.P2_has_type, getattr(AAT, '300379244')))
        new_graph.add((getattr(AAT, '300379244', CRM.P1_is_identified_by, Literal('years@en'))))
        new_graph.add((time_span_PID, CRM.P1_is_identified_by, Literal(obj)))
        new_graph.add((time_span_PID, CRM.P82_at_some_time_within, Literal(obj, datatype=XSD.gYear)))
        new_graph.add((time_span_PID, OWL.sameAs, wikidata_year))

    return new_graph

def create_event_triples(subject_PID, subj, pred):
    if pred == getattr(RRO, 'RP68.was_acquired'):
        event_PID = create_PID_from_triple('acquisition', subj)
    elif pred == getattr(RRO, 'RP72.was_produced'):
        event_PID = create_PID_from_triple('production', subj)
        event_type = CRM.E12_Production
        aat_event_id = getattr(AAT, '300404387')
        aat_event_type = Literal('creating (artistic activity)@en')
    elif pred == getattr(RRO, 'RP34.was_born_in'):
        event_PID = create_PID_from_triple('birth', subj)
    elif pred == getattr(RRO, 'RP4.died_in'):
        event_PID = create_PID_from_triple('death', subj)

    new_graph.add((getattr(NGO, subject_PID), CRM.P12_was_present_at, getattr(NGO, event_PID)))
    new_graph.add((getattr(NGO, event_PID), CRM.P2_has_type, event_type))
    new_graph.add((getattr(NGO, event_PID), CRM.P2_has_type, aat_event_id))
    new_graph.add((aat_event_id, CRM.P1_is_identified_by, aat_event_type))

    return new_graph
    
def map_object(old_graph, new_graph):
    for painting_id, _, _ in old_graph.triples((None, RDF.type, getattr(RRO,'RC12.Painting'))):
        for subj, pred, obj in old_graph.triples((painting_id, None, None)):
            subject_PID = generate_placeholder_PID(subj)
            new_graph = create_title_triples(subject_PID, subj, pred, obj)
            new_graph = create_medium_triples(subject_PID, subj, pred, obj)
            new_graph = create_collection_triples(subject_PID, subj, pred, obj)
            new_graph = create_dimension_triples(subject_PID, subj, pred, obj)
            new_graph = create_identifier_triples(subject_PID, pred, obj)
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_event_triples(subject_PID, subj, pred)
            #or would it be wiser to call map_event from inside this function? not sure if that makes everything
            #too messy or if it's workable

    return new_graph

'''
def map_event(old_graph, new_graph):
    for event_id, _, _ in old_graph.triples((None, getattr(RRO, 'RP72.was_produced'))):
        for subj, pred, obj in old_graph.triples((event_id, None, None)):
            subject_PID = generate_placeholder_PID(subj)

    return new_graph
'''

#new_graph = map_property(g, RDF.type, RDFS.subClassOf)
#query_graph(new_graph,None,RDFS.subClassOf,None)

#new_dataframe = sparql_query_pythonic('NG1171',csv_format=False)
#mapped_dataframe = map_property(new_dataframe,'NG1171','PID')
#print(mapped_dataframe)
#triples_to_csv(new_dataframe)

#generate_placeholder_PID('Fake Painting 3')

#query_graph(g, None, getattr(RRO,'RP34.has_title'), None)
#map_property(g, new_graph, getattr(RRO,'RP34.has_title'), getattr(CRM,'P102_Has_Title'))
#query_graph(g, None, None, getattr(RRI, 'RC239.Exhibition'))
#triples_to_csv(new_graph,'crm_mapping_draft')

new_mapping = map_object(g,new_graph)
#for x,y,z in new_mapping:
#    print(x,y,z)
triples_to_csv(new_mapping,'painting_triples')